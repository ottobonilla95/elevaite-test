'''
Functions pretaining to
    1. Upload Excel Workbook
    2. Process Excel File - Size, list of sheets, Dataframe creation, csv conversion
'''

from fastapi import UploadFile, File
import openpyxl
import os
import pandas as pd
import re


def convert_bytes_to_human_readable(size_in_bytes):
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size_in_bytes < 1024.0:
            return f"{size_in_bytes:.2f} {unit}"
        size_in_bytes /= 1024.0


async def upload_file(file: UploadFile = File(...)):
    try:

        #Get File Type
        file_type = file.content_type

        #Get File Size
        file.file.seek(0,2)
        size = file.file.tell()
        file.file.seek(0)
        file_size = convert_bytes_to_human_readable(size)

        #Get File Name
        file_name = file.filename

        #Save Excel file to local drive
        os.makedirs("data/Excel", exist_ok=True)
        with open(f"data/Excel/{file.filename}", "wb") as f:
            f.write(file.file.read())
        file_path = f"data/Excel/{file.filename}"
        #Excel/Workbook - get list of sheets
        sheets = []
        sheets_count = 0
        if file_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or file_type == "application/vnd.ms-excel":
            file_path = f"data/Excel/{file.filename}"
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheets = workbook.sheetnames
            sheets_count = len(sheets)
        return {"response": "Success", "file_size": file_size, "file_path": file_path, "file_name": file_name, "file_type" : file_type, "sheet": sheets, "sheets_count": sheets_count}
    except Exception as e:
        return {"response": "Error", "error_message": e}
    

def find_numeric_cells(sheet):
    
    # Initialize variables to store the first and last row numbers and column numbers
    first_row = None
    last_row = None
    first_col = None
    last_col = None

    # Iterate through rows and columns in the sheet
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for col_index, cell_value in enumerate(row, start=1):
            # Check if the cell value is numeric
            if isinstance(cell_value, (int, float)):
                # Update first_row if not set
                if first_row is None:
                    first_row = row_index

                # Update last_row continuously
                last_row = row_index

                # Update first_col if not set
                if first_col is None:
                    first_col = col_index

                # Update last_col continuously
                if last_col is None:
                    last_col = col_index
                if last_col < col_index:
                    last_col = col_index


def get_index(start_row, end_row, start_column, end_column, sheet):
    
    row_headers = []
    col_headers = []

    for row_index in range(start_row, end_row + 1):
        header = ""
        for col_index in range(0, start_column):
            try:
                cell_value = sheet.cell(row=row_index, column=col_index).value
                if header == "" and cell_value is not None:
                    header = re.sub(r'[^a-zA-Z0-9]', '', str(cell_value))
                elif header != "" and cell_value is not None:
                    header+="_"+re.sub(r'[^a-zA-Z0-9]', '', str(cell_value))
            except Exception:
                #print("Error:", str(e))
                print(".")
        
        row_headers.append(header)
    for col_index in range(start_column, end_column+1):
        header = ""
        for row_index in range(0, start_row):
            try:
                cell_value = sheet.cell(row=row_index, column=col_index).value
                
                if header == "" and cell_value is not None:
                    header = re.sub(r'[^a-zA-Z0-9]', '', str(cell_value))
                elif header != "" and cell_value is not None:
                    header += "_"+re.sub(r'[^a-zA-Z0-9]', '', str(cell_value))
            except Exception:
                #print("Error:", str(e))
                print(".")
        col_headers.append(header)

    return row_headers, col_headers

def create_dataframe_2(start_row, start_col, end_row, end_col, row_headers, col_headers,workbook_name,sheet):
    row_data = {}
    
    j = 0
    # Iterate through rows and columns in the specified range and populate the list
    for row_index in range(start_row, end_row + 1):
        i = 0
        for col_index in range(start_col, end_col + 1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            header = row_headers[j] + "_" + col_headers[i]
            row_data[header] = cell_value
            i += 1
        j += 1
        
            
    # Create an empty DataFrame
    #df = pd.DataFrame(row_data, index=False)
    df = pd.DataFrame([row_data])
   
    
    csv_file_dir = os.path.join(os.getcwd(), "data", "Output", workbook_name.split(".")[0])
    os.makedirs(csv_file_dir, exist_ok=True)
    csv_file = sheet.title + ".csv"
    csv_file_path = os.path.join(csv_file_dir, csv_file)
    mode = 'w'
    df.to_csv(csv_file_path, mode = mode)
    return csv_file_path

def  generate_csv_for_excel(workbook_path):


    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    workbook_name = workbook_path.split('/')[-1]
    sheets = workbook.sheetnames
    csv_sheets = []
    for sheet_name in sheets:
        sheet = workbook[sheet_name] 
        
        first_row, first_col, last_row, last_col = find_numeric_cells(sheet)
        row_headers, col_headers = get_index(first_row, last_row, first_col, last_col, sheet)
        csv_sheet = create_dataframe_2(first_row, first_col, last_row, last_col, row_headers, col_headers, workbook_name, sheet)
        csv_sheets.append(csv_sheet)
    return {"response" : "Success", "sheet_list" : csv_sheets}